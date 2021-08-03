import openpyxl
import json
from fpdf import FPDF,HTMLMixin
from PIL import Image
wb = openpyxl.load_workbook('dummy data.xlsx', data_only=True)
sh1 = wb['Sheet1']
row = sh1.max_row
column = sh1.max_column
arr = []
card = {}
res = ""

for x in range(3, row+1):
    if res == sh1.cell(x, 6).value:
        card['question'].append({
            "Question No.": str(sh1.cell(x, 14).value),
            "What you marked": str(sh1.cell(x, 15).value),
            "Correct Answer": str(sh1.cell(x, 16).value),
            "Outcome(Correct/Incorrect/Not Attempted)": str(sh1.cell(x, 17).value),
            "Score if correct": str(sh1.cell(x, 18).value),
            "Your score": str(sh1.cell(x, 19).value)
        })
    else:
        res = sh1.cell(x, 6).value
        arr.append(card)
        card = {
            "Round": str(sh1.cell(x, 2).value),
            "First Name": str(sh1.cell(x, 3).value),
            "Last Name": str(sh1.cell(x, 4).value),
            "Full Name": sh1.cell(x, 5).value,
            "Registration Number": str(sh1.cell(x, 6).value),
            "Grade": str(sh1.cell(x, 7).value),
            "Name of School": str(sh1.cell(x, 8).value),
            "Gender": sh1.cell(x, 9).value,
            "Date of Birth": str(sh1.cell(x, 10).value),
            "City of Residence": str(sh1.cell(x, 11).value),
            "Date and time of test": str(sh1.cell(x, 12).value),
            "Country of Residence": sh1.cell(x, 13).value,
            "question": [{
                "Question No.": str(sh1.cell(x, 14).value),
                "What you marked": str(sh1.cell(x, 15).value),
                "Correct Answer": str(sh1.cell(x, 16).value),
                "Outcome(Correct/Incorrect/Not Attempted)": str(sh1.cell(x, 17).value),
                "Score if correct": str(sh1.cell(x, 18).value),
                "Your score": str(sh1.cell(x, 19).value)
            }],
            "Final result": str(sh1.cell(x, 20).value)
        }
arr.append(card)
arr.pop(0)
# class PDF(FPDF,HTMLMixin):
#     pass

for x in arr:
    WIDTH = 210
    HEIGHT = 297
    document = FPDF( format="A4")
    document.add_page()
    # document.set_image_filter("DCTDecode")
    img = Image.open("img/"+x["First Name"]+" "+x["Last Name"]+".png")
    x_=img.size[0]
    y_=img.size[1]
    img = img.crop((0,0,x_,y_)).resize((100,100),resample=Image.NEAREST)
    document.image(img,x=document.epw-70,y=60)
    document.set_font('Helvetica', size=12)
    document.ln(60)
    document.cell(w=50,h=2,txt="Round - "+x["Round"],markdown=True)
    document.ln(10)
    document.cell(txt="Full Name - "+x['First Name']+" "+x["Last Name"])
    document.ln(10)
    document.cell(w=50,h=2,txt="Registration Number - "+x["Registration Number"])
    document.ln(10)
    document.cell(w=50,h=2,txt="Grade - "+x["Grade"])
    document.ln(10)
    document.cell(w=50,h=2,txt="Name Of School - "+x["Name of School"])
    document.ln(10)
    document.cell(txt="Gender - "+x["Gender"])
    document.ln(10)
    document.cell(txt="Date of Birth - "+x["Date of Birth"])
    document.ln(10)
    document.cell(txt="City of Residence - "+x["City of Residence"])
    document.ln(10)
    document.cell(txt="Date and time of test - "+x["Date and time of test"])
    document.ln(10)
    document.cell(txt="Country of Residence - "+x["Country of Residence"])
    document.ln(10)
    document.add_page()
    document.set_font("Times", size=10)
    line_height = document.font_size * 2.5
    col_width = document.epw / 6  # distribute content evenly
    c1=(card['question'][0].keys())
    for row in c1:
        document.multi_cell(col_width,line_height,row,border=1,ln=3, max_line_height=document.font_size)
    document.ln(line_height)
    for row in card['question']:
        for datum in row.values():
            document.multi_cell(col_width, line_height,str(datum), border=1, ln=3, max_line_height=document.font_size)
        document.ln(line_height)
    document.output(x["First Name"]+".pdf")
